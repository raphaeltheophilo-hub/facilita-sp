"""
utils.py — Importador Excel, gerador de PDF e utilitários (versão cloud)
Sistema Facilita SP · Diretoria de Aumento da Produtividade

Diferença da versão local:
- Documentos são salvos no PostgreSQL (BYTEA), não em disco
- Sem monitoramento de pasta (sem acesso a filesystem no Render)
"""
import io
from datetime import datetime

import pandas as pd
from fpdf import FPDF

import db


# ── Importação da Planilha ────────────────────────────────────────────────────

def import_from_bytes(file_bytes: bytes, filename: str) -> tuple[bool, str]:
    try:
        df = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name="Tabela de Municípios",
            dtype={"CÓDIGO IBGE": str},
        )
        df = df.dropna(subset=["CÓDIGO IBGE"])
        df["CÓDIGO IBGE"] = df["CÓDIGO IBGE"].astype(str).str.strip()
        df = df[df["CÓDIGO IBGE"].str.match(r"^\d{7}$")]
        rows = df.to_dict("records")
        if not rows:
            db.log_importacao(filename, 0, "erro", "Nenhuma linha valida encontrada.")
            return False, "Nenhuma linha valida encontrada na planilha."
        count = db.upsert_municipios(rows)
        db.log_importacao(filename, count, "sucesso")
        return True, f"{count} municipios importados/atualizados com sucesso."
    except Exception as exc:
        db.log_importacao(filename, 0, "erro", str(exc))
        return False, f"Erro na importacao: {exc}"


# ── Salvar documentos (no banco, sem disco) ───────────────────────────────────

def save_document(codigo_ibge: str, uploaded_file, descricao: str, enviado_por: str):
    """Lê os bytes do arquivo e salva direto no PostgreSQL como BYTEA."""
    conteudo = uploaded_file.getbuffer().tobytes()
    ext = uploaded_file.name.rsplit(".", 1)[-1].lower() if "." in uploaded_file.name else "bin"
    db.add_documento(
        codigo_ibge=codigo_ibge,
        nome_arquivo=uploaded_file.name,
        conteudo_bytes=conteudo,
        tipo_arquivo=ext,
        descricao=descricao,
        enviado_por=enviado_por,
    )


# ── Geração de Relatório PDF ──────────────────────────────────────────────────

AZUL      = (24, 95, 165)
AZUL_CLR  = (235, 243, 253)
CINZA     = (100, 100, 100)
CINZA_CLR = (245, 245, 245)
PRETO     = (30, 30, 30)
BRANCO    = (255, 255, 255)
VERDE     = (22, 163, 74)
VERMELHO  = (220, 38, 38)
OURO      = (212, 175, 55)
PRATA     = (168, 169, 173)
BRONZE    = (205, 127, 50)


class _PDF(FPDF):
    def header(self):
        self.set_fill_color(*AZUL)
        self.rect(0, 0, 210, 18, "F")
        self.set_font("Helvetica", "B", 11)
        self.set_text_color(*BRANCO)
        self.set_xy(8, 4)
        self.cell(130, 7, "FACILITA SP - Secretaria de Desenvolvimento Economico")
        self.set_font("Helvetica", "", 9)
        self.set_xy(148, 4)
        self.cell(55, 7, f"Emitido: {datetime.now().strftime('%d/%m/%Y %H:%M')}", align="R")
        self.set_text_color(*PRETO)
        self.ln(16)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(*CINZA)
        self.cell(0, 6, f"Pag. {self.page_no()} - Diretoria de Aumento da Produtividade / SDE-SP", align="C")

    def titulo_secao(self, texto: str):
        self.ln(3)
        self.set_fill_color(*AZUL)
        self.set_text_color(*BRANCO)
        self.set_font("Helvetica", "B", 10)
        self.cell(0, 7, f"  {texto}", fill=True)
        self.set_text_color(*PRETO)
        self.ln(9)

    def linha_indicador(self, label: str, valor: str, alternado: bool):
        self.set_fill_color(*(CINZA_CLR if alternado else BRANCO))
        sim = str(valor).upper() in ("SIM", "S")
        nao = str(valor).upper() in ("NAO", "N")
        self.set_font("Helvetica", "", 10)
        self.cell(110, 7, f"  {label}", fill=True, border="B")
        if sim:
            self.set_text_color(*VERDE); self.set_font("Helvetica", "B", 10)
        elif nao:
            self.set_text_color(*VERMELHO); self.set_font("Helvetica", "B", 10)
        self.cell(80, 7, str(valor or "-"), fill=True, border="B", align="C")
        self.set_text_color(*PRETO)
        self.ln()

    def cabecalho_tabela(self, colunas: list):
        self.set_fill_color(*AZUL_CLR)
        self.set_font("Helvetica", "B", 9)
        for texto, largura in colunas:
            self.cell(largura, 7, texto, fill=True, border=1, align="C")
        self.ln()

    def linha_tabela(self, valores: list, alternado: bool):
        self.set_fill_color(*(CINZA_CLR if alternado else BRANCO))
        self.set_font("Helvetica", "", 9)
        for texto, largura in valores:
            self.cell(largura, 6, str(texto or "")[:50], fill=True, border="B")
        self.ln()


def gerar_pdf_municipio(mun, historico, documentos) -> bytes:
    pdf = _PDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(*AZUL)
    pdf.cell(0, 10, str(mun["nome"]), ln=True)
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(*CINZA)
    pdf.cell(0, 6,
             f"IBGE: {mun['codigo_ibge']}   Lat: {mun['latitude']}   Lng: {mun['longitude']}   "
             f"Dados: {str(mun['data_atualizacao'] or '')[:10] or 'sem data'}", ln=True)
    pdf.set_text_color(*PRETO)
    pdf.ln(2)

    cores_selo = {"OURO": OURO, "PRATA": PRATA, "BRONZE": BRONZE}
    cor_selo = cores_selo.get(str(mun["selo_principal"] or ""), CINZA)
    pdf.set_fill_color(*cor_selo)
    pdf.set_text_color(*BRANCO)
    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(70, 9, f"Selo 2025: {str(mun['selo_2025'] or 'Sem Avaliacao')}", fill=True, align="C")
    pdf.set_text_color(*PRETO)
    pdf.ln(12)

    pdf.titulo_secao("INDICADORES SISTEMICOS")
    indicadores = [
        ("REDESIM - Adesao",       mun["adesao_redesim"]),
        ("REDESIM - Operando",     mun["operando_redesim"]),
        ("Adesao ao Facilita",     mun["adesao_facilita"]),
        ("Viabilidade Automatica", mun["viabilidade_automatica"]),
        ("Inscricao Municipal",    mun["inscricao_municipal"]),
        ("Tem Inovacao",           mun["tem_inovacao"]),
        ("Classificacao de Risco", mun["classificacao_risco"] or "-"),
        ("Provedor Tecnologico",   mun["provedor_tecnologico"] or "-"),
    ]
    for i, (lbl, val) in enumerate(indicadores):
        pdf.linha_indicador(lbl, val, i % 2 == 0)

    pdf.ln(4)
    pdf.titulo_secao(f"HISTORICO DE CONTATOS ({len(historico)} registro(s))")
    if historico:
        cols_h = [("Data",24),("Tipo",28),("Assunto",68),("Responsavel",40),("Contato",30)]
        pdf.cabecalho_tabela(cols_h)
        for i, r in enumerate(historico[:15]):
            pdf.linha_tabela([
                (str(r["data_contato"])[:10],    24),
                (str(r["tipo_contato"] or ""),   28),
                (str(r["assunto"] or ""),        68),
                (str(r["responsavel"] or ""),    40),
                (str(r["nome_contato"] or ""),   30),
            ], i % 2 == 0)
        if len(historico) > 15:
            pdf.set_font("Helvetica", "I", 8)
            pdf.set_text_color(*CINZA)
            pdf.cell(0, 5, f"  ... e mais {len(historico)-15} registro(s) no sistema.", ln=True)
            pdf.set_text_color(*PRETO)
    else:
        pdf.set_font("Helvetica", "I", 10)
        pdf.set_text_color(*CINZA)
        pdf.cell(0, 7, "  Nenhum contato registrado.", ln=True)
        pdf.set_text_color(*PRETO)

    pdf.ln(4)
    pdf.titulo_secao(f"DOCUMENTOS CADASTRADOS ({len(documentos)} arquivo(s))")
    if documentos:
        for i, doc in enumerate(documentos):
            pdf.set_fill_color(*(CINZA_CLR if i % 2 == 0 else BRANCO))
            tipo_tag = "[PDF]" if doc["tipo_arquivo"] == "pdf" else "[IMG]"
            pdf.set_font("Helvetica", "B", 9)
            pdf.cell(14, 6, tipo_tag, fill=True, border="B")
            pdf.set_font("Helvetica", "", 9)
            pdf.cell(118, 6, str(doc["nome_arquivo"] or "")[:58], fill=True, border="B")
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(58, 6,
                     f"{str(doc['data_upload'])[:10]}  {doc['enviado_por'] or '-'}",
                     fill=True, border="B")
            pdf.ln()
    else:
        pdf.set_font("Helvetica", "I", 10)
        pdf.set_text_color(*CINZA)
        pdf.cell(0, 7, "  Nenhum documento cadastrado.", ln=True)
        pdf.set_text_color(*PRETO)

    return bytes(pdf.output())


# ── Helpers ───────────────────────────────────────────────────────────────────

SELO_EMOJI   = {"OURO": "🥇", "PRATA": "🥈", "BRONZE": "🥉", "SEM AVALIAÇÃO": "⬜"}
STATUS_EMOJI = {"SIM": "✅", "NÃO": "❌"}
TIPOS_CONTATO = [
    "E-mail", "Telefone", "Reuniao presencial", "Reuniao virtual",
    "Visita tecnica", "Oficio", "WhatsApp", "Outro",
]


def fmt_sim_nao(valor: str) -> str:
    return STATUS_EMOJI.get(str(valor).upper(), "-") + " " + (valor or "-")


def fmt_selo(selo: str) -> str:
    base = str(selo or "").replace(" E INOVAÇÃO", "").strip()
    return SELO_EMOJI.get(base, "⬜") + " " + (selo or "-")


def import_interlocutores_from_bytes(file_bytes: bytes, filename: str) -> tuple[bool, str]:
    """Importa a planilha de interlocutores para o banco."""
    try:
        df = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name="Interlocutores",
            dtype={"CÓDIGO IBGE": str},
        )
        df = df.dropna(subset=["CÓDIGO IBGE"])
        df["CÓDIGO IBGE"] = df["CÓDIGO IBGE"].astype(str).str.strip().str.zfill(7)
        rows = df.to_dict("records")
        if not rows:
            return False, "Nenhuma linha válida encontrada."
        count = db.upsert_interlocutores(rows)
        db.log_importacao(filename, count, "sucesso", "Interlocutores")
        return True, f"{count} interlocutores importados com sucesso."
    except Exception as exc:
        db.log_importacao(filename, 0, "erro", str(exc))
        return False, f"Erro na importação: {exc}"


# ── Exportação do Histórico de Contatos ───────────────────────────────────────

def exportar_historico_excel(registros: list) -> bytes:
    """
    Gera Excel com registro completo por linha — notas sem truncamento,
    formatação por seção com cabeçalho destacado.
    """
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Histórico de Contatos"

    AZUL_HEX  = "185FA5"
    AZUL_CLR_HEX = "E6F1FB"
    CINZA_HEX = "F0F2F6"
    BORDA     = Side(style="thin", color="CCCCCC")

    def borda_fina():
        return Border(left=BORDA, right=BORDA, top=BORDA, bottom=BORDA)

    def borda_bottom():
        return Border(bottom=Side(style="medium", color="CCCCCC"))

    # ── Cabeçalho do relatório ────────────────────────────────────────────────
    from datetime import datetime as _dt
    ws.merge_cells("A1:I1")
    ws["A1"] = "Histórico de Contatos — Facilita SP"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=AZUL_HEX)
    ws["A1"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    ws["A2"] = (f"Gerado em: {_dt.now().strftime('%d/%m/%Y %H:%M')}   |   "
                f"{len(registros)} registro(s)")
    ws["A2"].font      = Font(size=9, color="666666", italic=True)
    ws["A2"].alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 16

    linha = 3  # linha atual

    for idx, r in enumerate(registros):
        linha += 1  # linha separadora entre registros
        cor_fundo = AZUL_CLR_HEX if idx % 2 == 0 else CINZA_HEX

        # ── Linha de título do registro ───────────────────────────────────────
        ws.merge_cells(f"A{linha}:I{linha}")
        municipio = str(r["nome_municipio"] or "")
        data      = str(r["data_contato"] or "")[:10]
        assunto   = str(r["assunto"] or "")
        ws[f"A{linha}"] = f"  {municipio}   ·   {data}   ·   {assunto}"
        ws[f"A{linha}"].font      = Font(bold=True, size=11, color="FFFFFF")
        ws[f"A{linha}"].fill      = PatternFill("solid", fgColor=AZUL_HEX)
        ws[f"A{linha}"].alignment = Alignment(vertical="center")
        ws.row_dimensions[linha].height = 20
        linha += 1

        # ── Campos em pares (label | valor | label | valor) ──────────────────
        def _linha_par(label1, val1, label2, val2):
            nonlocal linha
            for col, txt, bold, cor in [
                (1, label1, True,  AZUL_CLR_HEX),
                (2, val1,   False, "FFFFFF"),
                (5, label2, True,  AZUL_CLR_HEX),
                (6, val2,   False, "FFFFFF"),
            ]:
                c = ws.cell(row=linha, column=col, value=str(txt or ""))
                c.font      = Font(bold=bold, size=10,
                                   color=AZUL_HEX if bold else "333333")
                c.fill      = PatternFill("solid", fgColor=cor_fundo)
                c.alignment = Alignment(vertical="center",
                                        wrap_text=True, indent=1)
                c.border    = borda_fina()
            # merge das células de valor para dar espaço
            ws.merge_cells(f"B{linha}:D{linha}")
            ws.merge_cells(f"F{linha}:I{linha}")
            ws.row_dimensions[linha].height = 16
            linha += 1

        _linha_par("Município",         r["nome_municipio"],
                   "Tipo de Contato",   r["tipo_contato"])
        _linha_par("Nome do Contato",   r["nome_contato"],
                   "Cargo",             r["cargo_contato"])
        _linha_par("Responsável (DAP)", r["responsavel"],
                   "Registrado em",     str(r["criado_em"] or "")[:16])

        # Prazo (só se houver)
        if r.get("tem_prazo") and r.get("data_prazo"):
            prazo_val = str(r["data_prazo"])[:10]
            if r.get("prazo_ok"):
                prazo_val += "  ✓ Concluído"
            _linha_par("Data Limite (Prazo)", prazo_val, "", "")

        # ── Notas completas ───────────────────────────────────────────────────
        notas = str(r["notas"] or "").strip()
        if notas:
            # Rótulo
            lbl = ws.cell(row=linha, column=1, value="Notas")
            lbl.font      = Font(bold=True, size=10, color=AZUL_HEX)
            lbl.fill      = PatternFill("solid", fgColor=AZUL_CLR_HEX)
            lbl.alignment = Alignment(vertical="top", indent=1)
            lbl.border    = borda_fina()

            # Conteúdo da nota — sem limite de caracteres
            ws.merge_cells(f"B{linha}:I{linha}")
            cel = ws.cell(row=linha, column=2, value=notas)
            cel.font      = Font(size=10, color="333333")
            cel.fill      = PatternFill("solid", fgColor="FFFFFF")
            cel.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
            cel.border    = borda_fina()

            # Altura proporcional ao tamanho das notas
            n_linhas_estimadas = max(2, notas.count("\n") + 1,
                                     len(notas) // 90 + 1)
            ws.row_dimensions[linha].height = min(n_linhas_estimadas * 15, 200)
            linha += 1

        # Linha separadora
        for col in range(1, 10):
            c = ws.cell(row=linha, column=col, value="")
            c.border = borda_bottom()
        ws.row_dimensions[linha].height = 6
        linha += 1

    # ── Largura das colunas ───────────────────────────────────────────────────
    larguras = {1: 22, 2: 20, 3: 10, 4: 10, 5: 22, 6: 20, 7: 10, 8: 10, 9: 10}
    for col, w in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_historico_pdf(registros: list) -> bytes:
    """
    Gera PDF com uma ficha completa por registro — notas sem truncamento,
    layout retrato A4, uma ficha por registro com todos os campos.
    """
    from datetime import datetime as _dt

    class _HistPDF(_PDF):
        def header(self):
            self.set_fill_color(*AZUL)
            self.rect(0, 0, 210, 16, "F")
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(*BRANCO)
            self.set_xy(8, 3)
            self.cell(130, 6, "FACILITA SP - Secretaria de Desenvolvimento Economico")
            self.set_font("Helvetica", "", 8)
            self.set_xy(148, 3)
            self.cell(55, 6,
                      f"Emitido: {_dt.now().strftime('%d/%m/%Y %H:%M')}",
                      align="R")
            self.set_text_color(*PRETO)
            self.ln(14)

        def footer(self):
            self.set_y(-11)
            self.set_font("Helvetica", "I", 7)
            self.set_text_color(*CINZA)
            self.cell(0, 5,
                      f"Pag. {self.page_no()} - Historico de Contatos - Facilita SP",
                      align="C")

    pdf = _HistPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=14)
    pdf.add_page()

    # Título do relatório
    pdf.set_font("Helvetica", "B", 15)
    pdf.set_text_color(*AZUL)
    pdf.cell(0, 9, "Historico de Contatos - Facilita SP", ln=True)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*CINZA)
    pdf.cell(0, 5,
             f"Gerado em: {_dt.now().strftime('%d/%m/%Y %H:%M')}   |   "
             f"{len(registros)} registro(s)",
             ln=True)
    pdf.set_text_color(*PRETO)
    pdf.ln(4)

    LARGURA = 190  # largura útil em mm

    def _campo(label: str, valor: str, w_label=40, w_valor=None):
        """Imprime um par label/valor em linha única."""
        if w_valor is None:
            w_valor = LARGURA - w_label
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(*AZUL_CLR)
        pdf.set_text_color(*AZUL)
        pdf.cell(w_label, 6, f"  {label}", fill=True, border="B")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_fill_color(*BRANCO)
        pdf.set_text_color(*PRETO)
        pdf.cell(w_valor, 6, f"  {str(valor or '-')}", fill=True, border="B")
        pdf.ln()

    def _campo_duplo(l1, v1, l2, v2, w_label=40):
        """Dois campos lado a lado na mesma linha."""
        metade = LARGURA // 2
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(*AZUL_CLR)
        pdf.set_text_color(*AZUL)
        pdf.cell(w_label, 6, f"  {l1}", fill=True, border="B")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_fill_color(*BRANCO)
        pdf.set_text_color(*PRETO)
        pdf.cell(metade - w_label, 6, f"  {str(v1 or '-')}", fill=True, border="B")
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(*AZUL_CLR)
        pdf.set_text_color(*AZUL)
        pdf.cell(w_label, 6, f"  {l2}", fill=True, border="B")
        pdf.set_font("Helvetica", "", 9)
        pdf.set_fill_color(*BRANCO)
        pdf.set_text_color(*PRETO)
        pdf.cell(metade - w_label, 6, f"  {str(v2 or '-')}", fill=True, border="B")
        pdf.ln()

    def _notas_pdf(texto: str):
        """Imprime o bloco de notas com quebra de linha automática — sem limite."""
        if not texto or not texto.strip():
            return
        # Rótulo
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_fill_color(*AZUL_CLR)
        pdf.set_text_color(*AZUL)
        pdf.cell(LARGURA, 6, "  Notas", fill=True, border="B")
        pdf.ln()
        # Conteúdo com multi_cell para quebra automática de linha
        pdf.set_font("Helvetica", "", 9)
        pdf.set_fill_color(250, 250, 252)
        pdf.set_text_color(*PRETO)
        pdf.set_x(pdf.l_margin)
        pdf.multi_cell(
            LARGURA, 5, texto.strip(),
            border=0, fill=True, align="L",
        )

    for idx, r in enumerate(registros):
        # Verifica espaço — ficha mínima ocupa ~45mm; se não couber, nova página
        if pdf.get_y() > 240:
            pdf.add_page()

        cor_barra = AZUL if idx % 2 == 0 else (30, 70, 130)

        # ── Barra de título do registro ───────────────────────────────────────
        pdf.set_fill_color(*cor_barra)
        pdf.set_text_color(*BRANCO)
        pdf.set_font("Helvetica", "B", 10)
        municipio = str(r["nome_municipio"] or "")
        data      = str(r["data_contato"] or "")[:10]
        assunto   = str(r["assunto"] or "")
        pdf.cell(LARGURA, 7,
                 f"  {municipio}   |   {data}   |   {assunto}",
                 fill=True)
        pdf.set_text_color(*PRETO)
        pdf.ln(8)

        # ── Campos estruturados ───────────────────────────────────────────────
        _campo_duplo("Tipo de Contato", r["tipo_contato"],
                     "Responsavel (DAP)", r["responsavel"])
        _campo_duplo("Nome do Contato",  r["nome_contato"],
                     "Cargo",            r["cargo_contato"])
        _campo_duplo("Registrado em",
                     str(r["criado_em"] or "")[:16],
                     "Data do Contato",   str(r["data_contato"] or "")[:10])

        # Prazo
        if r.get("tem_prazo") and r.get("data_prazo"):
            prazo_val = str(r["data_prazo"])[:10]
            status    = "Concluido" if r.get("prazo_ok") else "Em aberto"
            _campo_duplo("Data Limite (Prazo)", prazo_val,
                         "Status do Prazo", status)

        # Notas completas (sem truncamento)
        _notas_pdf(str(r["notas"] or ""))

        # Espaçamento entre fichas
        pdf.ln(5)
        pdf.set_draw_color(180, 180, 180)
        pdf.line(pdf.l_margin, pdf.get_y(),
                 pdf.l_margin + LARGURA, pdf.get_y())
        pdf.ln(4)

    return bytes(pdf.output())

